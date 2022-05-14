import shutil
from django.shortcuts import render,HttpResponse
from django.http import HttpResponse,HttpResponseRedirect
from requests import request
from .models import FilesUpload,MemberDetails
from django.conf import settings
import openpyxl
import os
# from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from django.core.exceptions import PermissionDenied
from loanapp.decorators import check_session


# Create your views here.
# @check_session
def home_view(request):
    if request.method=="POST":
        shutil.rmtree(os.path.join(settings.BASE_DIR,'media')) #removes the directory media and all the files in it.
        filename=request.FILES["file"]
        print("filename",filename)
        document=FilesUpload.objects.create(file=filename)
        document.save()  #saves the file to media directory

        path=os.path.join(settings.BASE_DIR,'media/%s' %(filename))
        
        wb_obj=openpyxl.load_workbook(path,data_only=True)
        sheet_obj = wb_obj.active

        # Delete all data from MemberDetails table
        MemberDetails.objects.all().delete()
        # Delete all data from MemberDetails table

        column_nos=sheet_obj.max_column
        row_nos=sheet_obj.max_row
        insertFlag=1

        for r in range(row_nos):
            dict_car={ }
            insertFlag=1
            for c in range(column_nos):
                if r!=0:
                    cell_obj = sheet_obj.cell(row = r+1, column = c+1)
                    if c+1==2:
                        if not cell_obj.value:
                            insertFlag=0
                            break
                        dict_car['member_number']=cell_obj.value
                    if c+1==3:
                        dict_car['member_type']=cell_obj.value
                    if c+1==4:
                        dict_car['name']=cell_obj.value
                    if c+1==5:
                        dict_car['sex']=cell_obj.value
                    if c+1==6:
                        dict_car['father_husband_name']=cell_obj.value
                    if c+1==7:
                        dict_car['house_name']=cell_obj.value
                    if c+1==8:
                        dict_car['kara']=cell_obj.value
                    if c+1==9:
                        dict_car['post']=cell_obj.value
                    if c+1==10:
                        dict_car['pin']=cell_obj.value
                    if c+1==11:
                        dict_car['village']=cell_obj.value
                    if c+1==12:
                        dict_car['taluk']=cell_obj.value
                    if c+1==13:
                        dict_car['panchayath']=cell_obj.value
                    if c+1==14:
                        dict_car['ward']=cell_obj.value
                    if c+1==15:
                        dict_car['mobile_number']=cell_obj.value
                    if c+1==16:
                        dict_car['dob']=cell_obj.value
                    if c+1==17:
                        dict_car['age']=cell_obj.value
                    if c+1==18:
                        dict_car['aadhar']=cell_obj.value
                    
            if r!=0 and insertFlag!=0:
                member_data=MemberDetails.objects.create(
                    member_number=dict_car['member_number'],member_type=dict_car['member_type'],name=dict_car['name'],sex=dict_car['sex'],
                    father_husband_name=dict_car['father_husband_name'],house_name=dict_car['house_name'],kara=dict_car['kara'],post=dict_car['post'],pin=dict_car['pin'],
                    village=dict_car['village'],taluk=dict_car['taluk'],panchayath=dict_car['panchayath'],ward=dict_car['ward'],
                    mobile_number=dict_car['mobile_number'],dob=dict_car['dob'],age=dict_car['age'],aadhar=dict_car['aadhar'],
                )   
                print("Member Data",member_data)    


        return render(request,"loan-agreement.html")
    session=request.session
    return render(request,"home-view.html")
    # if request.session.has_key('username'):
    #     return render(request,"home-view.html")
    # else:
    #     raise PermissionDenied


def RenderHTML(request):
    print(request.POST.get('txtVaypa'))
    jamyam1_membernumber=request.POST.get('txtJamyam1')
    jamyam2_membernumber=request.POST.get('txtJamyam2')
    jamyam3_membernumber=request.POST.get('txtJamyam3')

    jamyam1_name=""
    jamyam1_taluk=""
    jamyam1_village=""
    jamyam1_kara=""
    jamyam1_house=""
    jamyam1_father_husband=""
    jamyam1_age=""

    getJamyam1Details=MemberDetails.objects.filter(member_number=jamyam1_membernumber).all()
    for mem in getJamyam1Details:
        jamyam1_name=mem.name
        jamyam1_taluk=mem.taluk
        jamyam1_village=mem.village
        jamyam1_kara=mem.kara
        jamyam1_house=mem.house_name
        jamyam1_father_husband=mem.father_husband_name
        jamyam1_age=mem.age

    loan_data={
        "vaypanumber":request.POST.get('txtVaypa') if request.POST.get('txtVaypa')!="" else "" ,
        "thuka":request.POST.get('txtThuka') if request.POST.get('txtThuka')!="" else "",
        "jamyam1":request.POST.get('txtJamyam1'),
        "jamyam1_name":jamyam1_name,
        "jamyam1_taluk":jamyam1_taluk,
        "jamyam1_village":jamyam1_village,
        "jamyam1_kara":jamyam1_kara,
        "jamyam1_house":jamyam1_house,
        "jamyam1_father_husband":jamyam1_father_husband,
        "jamyam1_age":jamyam1_age,


        "jamyam2":request.POST.get('txtJamyam2'),
        "jamyam3":request.POST.get('txtJamyam3')
    }

    template=get_template("loan-agreement.html")
    html=template.render(loan_data)
    return HttpResponse(html)